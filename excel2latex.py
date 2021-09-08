"""Converts excel file sheet into latex table with same background color.
The idea is taken from: https://tex.stackexchange.com/a/572444/251678
"""
import openpyxl
from openpyxl import load_workbook
import pandas as pd
import numpy as np

def define_colors(sheet):
    """Extracts background colors from cell of excel sheet.

    Args:
        excel_file (str): filename of excel file
        sheet_name (str): sheet name in excel file

    Returns:
        [set]: All colors (Hex and rgb) from excel sheet
        [numpy array]: Hex colors as matrix like excel sheet
    """
    colors = []
    max_row = sheet.max_row
    max_col = sheet.max_column

    for row in range(1, max_row +1):
        for c in range(1, max_col+1):
            cell_color = sheet.cell(row, c).fill.start_color.index
            hex_color = cell_color[2:]
            r, g, b = tuple(round(int(cell_color[i+2:i+4], 16)* 100/255)/100 for i in (0, 2, 4))
            colors.append((hex_color, r,g,b))

    color_matrix = np.array([el[0] for el in colors]).reshape(max_row , max_col)

    return set(colors), color_matrix


def create_latex_table(sheet):
    df = pd.DataFrame(list(sheet.values)[1:], columns=list(sheet.values)[0])

    colors, color_matrix = define_colors(sheet)

    val_color = np.dstack((df.values,color_matrix[1:]))  # 3d array with ["value", "color"]

    # convert 3d to 2d array with ["value_color"]
    val_color_merged = [["" for j in range(val_color.shape[1])] for i in range(val_color.shape[0])]  # ini
    for i,row in enumerate(val_color):
        for j,tup in enumerate(row):
            val_color_merged[i][j] += "_".join([str(el) for el in tup])

    df_colors = pd.DataFrame(val_color_merged, columns=df.columns)

    column_format = "|"+"l|"*len(df_colors.columns)

    def formater(x):
        split = x.split('_')
        age = split[0]
        colour = split[1]
        return '\cellcolor{'+str(colour)+'}{'+str(age)+'}'

    # color definitions for tex file header
    with open("color_definitions.txt", "w") as f:
        for color in colors:
            line = "\definecolor{%s}{rgb}{%s,%s,%s}" % color
            f.write(line+"\r\n")

    # latex table
    with open("latex_table.txt", "w") as f:
        print(df_colors.to_latex(index=False, column_format=column_format, formatters=[formater for el in range(len(df_colors.columns))], escape=False), file=f)


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Converts excel sheet to latex table with same background color.')
    parser.add_argument('filename', type=str, help='Excel filename')
    parser.add_argument('sheetname', type=str, help='Excel sheet name')
   
    args = parser.parse_args()

    excel_file, sheet_name = args.filename, args.sheetname

    workbook = load_workbook(excel_file, data_only = True)
    sheet = workbook[sheet_name]

    create_latex_table(sheet)